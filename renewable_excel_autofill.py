#!/usr/bin/env python3
"""
Renewable Project Tender → Excel Autofill
-----------------------------------------
Reads a folder of tender documents (PDF/DOCX/TXT/XLS/XLSX), extracts
key tags using regex rules, and writes rows into a Master sheet that
matches your existing schema (auto-derived from 8.1 Lakhwar_Master_v8_MLready_v2.xlsx).

Usage:
  python renewable_excel_autofill.py --docs ./input_docs \\
      --out ./renewable_output.xlsx \\
      --columns ./renewable_columns_master.json \\
      --tags ./tags_renewable.yaml \\
      [--template /path/to/your_template.xlsx]

Dependencies:
  pip install pandas openpyxl python-docx PyPDF2 pyyaml
  (Optional: pdfminer.six pdfplumber python-magic)

Notes:
  - This is a starter. Extend TAG_RULES and MAPPINGS as needed.
  - The script tries to be tolerant: missing libs are handled gracefully.

"""

import os, re, json, sys, argparse, logging, io
from pathlib import Path
from datetime import datetime

# Optional deps
try:
    import yaml
except Exception:
    yaml = None

import pandas as pd
from openpyxl import Workbook, load_workbook

# Best-effort PDF and DOCX readers
def read_pdf_text(path: Path) -> str:
    # Try PyPDF2
    try:
        import PyPDF2
        text_parts = []
        with open(path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                try:
                    text_parts.append(page.extract_text() or "")
                except Exception:
                    continue
        return "\n".join(text_parts)
    except Exception as e:
        logging.warning(f"PyPDF2 failed on {path.name}: {e}")
    # Try pdfminer.six
    try:
        from pdfminer.high_level import extract_text
        return extract_text(str(path))
    except Exception as e:
        logging.warning(f"pdfminer failed on {path.name}: {e}")
    # Try pdfplumber
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                text_parts.append(page.extract_text() or "")
        return "\n".join(text_parts)
    except Exception as e:
        logging.warning(f"pdfplumber failed on {path.name}: {e}")
    return ""

def read_docx_text(path: Path) -> str:
    try:
        import docx
        doc = docx.Document(str(path))
        return "\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        logging.warning(f"python-docx failed on {path.name}: {e}")
        return ""

def read_xl_text(path: Path) -> str:
    try:
        xls = pd.ExcelFile(path)
        parts = []
        for sheet in xls.sheet_names:
            try:
                df = pd.read_excel(path, sheet_name=sheet).astype(str)
                parts.append(df.to_csv(index=False))
            except Exception:
                continue
        return "\n".join(parts)
    except Exception as e:
        logging.warning(f"Excel read failed on {path.name}: {e}")
        return ""

def read_any_text(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return read_pdf_text(path)
    if ext in {".docx"}:
        return read_docx_text(path)
    if ext in {".xlsx", ".xls"}:
        return read_xl_text(path)
    if ext in {".txt", ".csv"}:
        try:
            return Path(path).read_text(encoding="utf-8", errors="ignore")
        except Exception:
            return Path(path).read_text(encoding="latin-1", errors="ignore")
    # fallback: try read as text
    try:
        return Path(path).read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""

def load_columns(columns_path: Path) -> dict:
    meta = json.loads(Path(columns_path).read_text(encoding="utf-8"))
    columns = meta.get("columns", [])
    sheet = meta.get("detected_master_sheet", "Master")
    return sheet, columns

def load_tags(tags_path: Path) -> dict:
    raw = Path(tags_path).read_text(encoding="utf-8")
    if yaml:
        return yaml.safe_load(raw)
    # Minimal YAML-ish parser fallback: extremely simple, supports this file.
    rules = {}
    current_key = None
    for line in raw.splitlines():
        if not line.strip() or line.strip().startswith("#"):
            continue
        if not line.startswith(" "):  # new key
            key = line.split(":", 1)[0].strip()
            rules[key] = []
            current_key = key
        else:
            m = re.search(r"-\s*'(.*)'\s*$", line.strip())
            if m and current_key:
                rules[current_key].append(m.group(1))
    return rules

def normalize_space(s: str) -> str:
    return re.sub(r"[\u00A0\t ]+", " ", s.replace("\r", "\n")).strip()

def extract_tags(text: str, tag_rules: dict) -> dict:
    text = normalize_space(text)
    out = {}
    for tag, patterns in tag_rules.items():
        for pat in patterns:
            try:
                m = re.search(pat, text, flags=re.IGNORECASE|re.MULTILINE)
            except re.error as e:
                logging.warning(f"Bad regex for {tag}: {e}")
                continue
            if m:
                val = m.groupdict().get("value", m.group(0))
                out[tag] = normalize_space(val)
                break
    return out

def build_row(columns: list, tag_values: dict) -> list:
    row = []
    for col in columns:
        # Simple mapping: if a tag of same snake_case exists, use it; else blank
        key = re.sub(r"[^A-Za-z0-9]+", "_", col).lower().strip("_")
        # Prefer exact key; else try conservative fuzzy matches
        if key in tag_values:
            row.append(tag_values[key])
        else:
            # small heuristic mappings
            aliases = [
                ("project_name", ["project", "tender_name", "name_of_work"]),
                ("project_capacity_mw", ["capacity_mw", "capacity"]),
                ("storage_capacity_mwh", ["bess_mwh", "storage_mwh"]),
                ("bid_submission_deadline", ["submission_deadline", "bid_due_date"]),
                ("emd_amount_rs", ["emd", "earnest_money"]),
                ("pbg_percent_or_amount", ["pbg", "performance_bg"]),
                ("completion_time_months", ["completion_time", "time_for_completion"]),
                ("price_cap_rs_per_kwh", ["tariff_cap", "ceiling_tariff"]),
                ("interconnection_voltage_kv", ["grid_voltage_kv", "voltage_kv"]),
            ]
            filled = ""
            for tag, aka in aliases:
                if key in aka and tag in tag_values:
                    filled = tag_values[tag]
                    break
            row.append(filled)
    return row

def ensure_workbook(out_path: Path, template: Path|None, master_sheet: str, columns: list):
    if template and template.exists():
        wb = load_workbook(str(template))
        if master_sheet not in wb.sheetnames:
            ws = wb.create_sheet(master_sheet)
            ws.append(columns)
        else:
            ws = wb[master_sheet]
            if ws.max_row < 1:
                ws.append(columns)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = master_sheet
        ws.append(columns)
    wb.save(str(out_path))

def append_rows(out_path: Path, master_sheet: str, rows: list[list]):
    wb = load_workbook(str(out_path))
    ws = wb[master_sheet]
    for r in rows:
        ws.append(r)
    wb.save(str(out_path))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--docs", required=True, help="Folder with input tender files")
    ap.add_argument("--out", required=True, help="Output Excel path")
    ap.add_argument("--columns", required=True, help="JSON file with master columns/schema")
    ap.add_argument("--tags", required=True, help="YAML file with tag regex rules")
    ap.add_argument("--template", default="", help="Optional template workbook to start from")
    ap.add_argument("--log", default="INFO", help="Log level")
    args = ap.parse_args()

    logging.basicConfig(level=getattr(logging, args.log.upper(), logging.INFO),
                        format="%(levelname)s: %(message)s")

    docs_dir = Path(args.docs)
    out_path = Path(args.out)
    template = Path(args.template) if args.template else None

    master_sheet, columns = load_columns(Path(args.columns))
    tag_rules = load_tags(Path(args.tags))

    logging.info(f"Master sheet: {master_sheet} ({len(columns)} columns)")
    ensure_workbook(out_path, template, master_sheet, columns)

    rows_to_write = []
    supported = {".pdf",".docx",".txt",".csv",".xlsx",".xls"}
    for p in sorted(docs_dir.glob("**/*")):
        if not p.is_file():
            continue
        if p.suffix.lower() not in supported:
            continue
        logging.info(f"Parsing {p.name}")
        text = read_any_text(p)
        if not text.strip():
            logging.warning(f"No text extracted from {p.name}")
            continue
        tags = extract_tags(text, tag_rules)
        # Always attach filename + timestamp if those columns exist
        tags.setdefault("source_file", p.name)
        tags.setdefault("ingested_at", datetime.now().isoformat(timespec="seconds"))
        row = build_row(columns, tags)
        rows_to_write.append(row)

    if rows_to_write:
        append_rows(out_path, master_sheet, rows_to_write)
        logging.info(f"Wrote {len(rows_to_write)} rows to {out_path}")
    else:
        logging.warning("No rows to write — did not find supported files or matches.")

if __name__ == "__main__":
    main()